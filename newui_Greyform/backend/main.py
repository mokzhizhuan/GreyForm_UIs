from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import Response
from pydantic import BaseModel
from typing import List, Dict, Any, Optional
import json
import shlex
import io
import os
import openpyxl

from backend.marking_controller import app as marking_subapp
from backend.ssh_client import run_command, read_remote_file

# ============================================================
# 🌐 Main FastAPI Application (NO ROS)
# ============================================================
app = FastAPI(title="Main API (no ROS)")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)
app.mount("/marking", marking_subapp)


# ============================================================
# Serve PBU-specific images (stage/wall diagrams, placement
# photos) that live on the Linux PC alongside the Excel file —
# these can't be bundled as static frontend assets since they're
# different per PBU folder.
# ============================================================
IMAGE_MEDIA_TYPES = {
    "png": "image/png",
    "jpg": "image/jpeg",
    "jpeg": "image/jpeg",
}


@app.get("/get_image")
def get_image(folder: str, filename: str):
    # basename() strips any directory traversal attempt from filename
    safe_filename = os.path.basename(filename)
    remote_path = f"{folder.rstrip('/')}/{safe_filename}"
    try:
        image_bytes = read_remote_file(remote_path)
    except Exception as e:
        raise HTTPException(status_code=404, detail=f"Image not found: {safe_filename} ({e})")

    ext = safe_filename.rsplit(".", 1)[-1].lower() if "." in safe_filename else ""
    media_type = IMAGE_MEDIA_TYPES.get(ext, "application/octet-stream")
    return Response(content=image_bytes, media_type=media_type)


# ============================================================
# Robot endpoint (NO ROS NEEDED)
# ============================================================
@app.get("/getdirectory")
def getdirectory():
    cmd = (
        "python3 /home/winsys/pbu_marking_ros/directorysearch.py "
        "--directory /home/winsys/pbu_marking_ros/pbu_data/"
    )
    lines, code = run_command(cmd)
    if code != 0:
        raise HTTPException(
            status_code=500,
            detail=f"getdirectory failed (code {code})",
        )
    return {"ok": True, "data": lines}


@app.get("/jointtarget/connection")
def jointtarget_connection():
    cmd = (
        "python3 /home/winsys/pbu_marking_ros/homeposcheck.py "
        "--file /home/winsys/pbu_marking_ros/pbu_data/mockup/poses.json "
        "--target outside"
    )
    lines, code = run_command(cmd)
    if code != 0:
        raise HTTPException(
            status_code=500,
            detail=f"jointtarget_connection failed (code {code})",
        )
    return {"ok": True, "data": lines}


# ============================================================
# Read Directory (SSH)
# ============================================================
@app.post("/read_directory")
def read_directory():
    lines, code = run_command("ls /home")
    if code != 0:
        raise HTTPException(
            status_code=500,
            detail=f"read_directory failed (code {code})",
        )
    return {"ok": True, "data": lines}


# ============================================================
#  File Execute → return walls + max wall number
# ============================================================
class FileExecBody(BaseModel):
    directory: str
    folder: Optional[str] = None
    excelfile: Optional[str] = None


class WallInfo(BaseModel):
    wall: str
    count: int
    rows: List[Dict[str, Any]]


class FileExecuBody(BaseModel):
    folder: str  # user clicked folder (e.g. .../mockup/test_tmp)


# Values in the Status column that count as "this row is done".
# Extend this set if the marking script ever writes a different marker.
DONE_STATUS_VALUES = {"done"}


def determine_current_stage(excel_bytes: bytes) -> Dict[str, Any]:
    """
    Reads Stage 1 / Stage 2 / Stage 3 sheets directly from the Excel
    file and figures out which stage the operator should start (or
    resume) marking from.

    A stage counts as COMPLETE only if every row's Status cell is a
    "done" marker. The first incomplete stage (in order 1, 2, 3) is
    returned as the stage to start from. If all three are complete,
    currentStage is None.

    NOTE: this treats a stage as all-or-nothing — it does not currently
    distinguish "some walls in this stage are done, others aren't".
    If partial-stage resume (e.g. skip only the rows already marked)
    is needed, this is the place to also return the specific incomplete
    row indices per stage.
    """
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)
    stage_summary: Dict[int, Dict[str, Any]] = {}
    current_stage = None

    for stage_num in (1, 2, 3):
        sheet_name = f"Stage {stage_num}"
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        header = [c.value for c in ws[1]]
        try:
            status_col = header.index("Status") + 1
        except ValueError:
            # No Status column on this sheet — treat as not our concern
            continue

        total_rows = 0
        done_rows = 0
        for row in ws.iter_rows(min_row=2):
            status_cell = row[status_col - 1].value
            # Skip fully blank rows (no data at all on this row)
            if all(c.value is None for c in row):
                continue
            total_rows += 1
            status_str = str(status_cell).strip().lower() if status_cell is not None else ""
            if status_str in DONE_STATUS_VALUES:
                done_rows += 1

        is_complete = total_rows > 0 and done_rows == total_rows
        stage_summary[stage_num] = {
            "totalRows": total_rows,
            "doneRows": done_rows,
            "isComplete": is_complete,
        }

        if current_stage is None and not is_complete:
            current_stage = stage_num

    return {
        "currentStage": current_stage,  # None means all stages complete
        "stages": stage_summary,
    }


@app.post("/file_execute_data")
def file_execute_data(body: FileExecuBody):
    try:
        root_dir = body.folder
        # ------------------------------------------------------------
        # 1️⃣ Find the most recently modified .xlsx in the folder and
        #    output its ABSOLUTE path.
        #
        #    NOTE: the Excel filename is whatever the operator typed
        #    into the "Output Name" field of the Greyform IFC Extractor
        #    (e.g. "output.xlsx", "TERRAHL2.xlsx", anything) — it is
        #    NOT a fixed name, so we can't search for one exact filename.
        #    Picking the newest .xlsx assumes the most recent extractor
        #    run is the one you want to use for this folder.
        # ------------------------------------------------------------
        find_cmd = (
            f"cd '{root_dir}' && "
            f"find . -type d -iname '*_out' -prune -o "
            f"-type f -iname '*.xlsx' -printf '%T@ %p\\n' "
            f"| sort -rn | head -n1 | cut -d' ' -f2- | xargs -r realpath"
        )
        find_lines, _ = run_command(find_cmd)
        excel_path = find_lines[0].strip() if find_lines else ""
        if not excel_path:
            return {
                "ok": False,
                "returncode": 0,
                "error": "No .xlsx file found in this folder",
                "data": [],
            }
        print("USING EXCEL:", excel_path)

        # basename of the excel file without extension — this is the
        # prefix the IFC Extractor uses for placement photos, e.g.
        # "output.xlsx" -> output_name "output" -> "output_pos1.png"
        excel_basename = os.path.basename(excel_path)
        output_name = os.path.splitext(excel_basename)[0]

        # ------------------------------------------------------------
        # 1.5️⃣ Read the Excel file directly to determine which stage
        #    the operator should start from.
        # ------------------------------------------------------------
        try:
            excel_bytes = read_remote_file(excel_path)
            stage_info = determine_current_stage(excel_bytes)
        except Exception as e:
            # Don't hard-fail the whole endpoint if stage detection has
            # a problem — surface it, but let wall detection still work.
            stage_info = {"currentStage": None, "stages": {}, "stageError": str(e)}

        current_stage = stage_info.get("currentStage")
        stage_error = stage_info.get("stageError")

        if current_stage is None and not stage_error:
            # Every stage's Status column is already fully "done" —
            # nothing left to mark in this file.
            return {
                "ok": False,
                "returncode": 0,
                "error": "All stages are already marked as done in this Excel file — nothing left to mark.",
                "data": [],
                **stage_info,
            }

        # If stage detection itself failed for some reason, fall back to
        # processing all stages rather than silently blocking the flow —
        # the stageError field still reports what went wrong.
        stages_arg = str(current_stage) if current_stage is not None else "1,2,3"

        # ------------------------------------------------------------
        # 2️⃣ Run detectwalls.py, scoped to only the stage we determined
        #    still needs marking.
        # ------------------------------------------------------------
        run_cmd = (
            f"python3 /home/winsys/pbu_marking_ros/detectwalls.py "
            f"--filename {shlex.quote(excel_path)} --stages {shlex.quote(stages_arg)}"
        )
        lines, code = run_command(run_cmd)
        if code != 0:
            return {
                "ok": False,
                "returncode": code,
                "error": "\n".join(lines).strip() or f"detectwalls.py exited with code {code}",
                "data": lines,
            }
        return {
            "ok": True,
            "returncode": code,
            "data": lines,
            "outputName": output_name,
            **stage_info,
        }
    except Exception as e:
        return {"ok": False, "error": str(e), "data": []}


class CombineRequest(BaseModel):
    folder: str  # full path: /home/ros_user/pbu_data/mockup/PBU_TERRAHL2_out


@app.post("/combine_walls")
def combine_walls(req: CombineRequest):
    remote_cmd = f"python3 /home/winsys/combine_wall_excels.py '{req.folder}'"
    try:
        lines, code = run_command(remote_cmd, timeout=60)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"SSH call failed: {str(e)}")

    if code != 0:
        raise HTTPException(
            status_code=500, detail=f"SSH error: {chr(10).join(lines).strip()}"
        )

    # Parse the JSON return from combine_wall_excels.py
    try:
        response_data = json.loads("\n".join(lines).strip())
    except json.JSONDecodeError:
        raise HTTPException(
            status_code=500,
            detail=f"Invalid JSON returned from remote script: {lines}",
        )
    return response_data
